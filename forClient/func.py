import sqlite3
from datetime import datetime
import time
import random
import string
import asyncio
from time import time
from aiogram.types import KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from aiogram.types.web_app_info import WebAppInfo
import openpyxl
from num2words import num2words

SHOP_ID = 312763
SHOP_API_TOKEN = 'live_JRW8BqO6Dj5tEWQ_MV68XewEaT9Xfs32NTxaOXWi39A'
date_now = int(time())
dt_object = datetime.fromtimestamp(date_now)


def make_pay_inline_key(url, payment):
    button = InlineKeyboardButton('Купить', web_app=WebAppInfo(
        url=url))
    check_button = InlineKeyboardButton(text='Проверить', callback_data='CheckPayStatus')
    keyboard = InlineKeyboardMarkup().add(button).add(check_button)

    return [keyboard, check_button]


def is_registered(user_id):
    conn = sqlite3.connect('db.db')
    cursor = conn.cursor()
    result = cursor.execute("SELECT user_tgname, user_name, user_phone FROM info WHERE user_id = ?",
                            [user_id]).fetchone()
    conn.commit()
    conn.close()
    if result is None:
        return False
    else:
        tgname, name, phone = result
        if tgname is None or name is None or phone is None:
            return False
        else:
            return True


def add_time(user_id, add_time):
    date_now = int(time())
    conn = sqlite3.connect('db.db')
    cursor = conn.cursor()
    result = cursor.execute("SELECT sub_time FROM info WHERE user_id=?", [user_id]).fetchone()
    conn.commit()
    sub_time = result[0]
    sub_time = sub_time + add_time
    cursor.execute("UPDATE info SET sub_time = ?, date_get_sub = ? WHERE user_id=?", [sub_time, date_now, user_id])
    conn.commit()
    cursor.close()
    conn.close()


def generate_keys(num_keys, time):
    conn = sqlite3.connect('db.db')
    cursor = conn.cursor()
    rusult = []
    for i in range(num_keys):
        key = ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))
        cursor.execute("INSERT INTO keys (key, time, user, status) VALUES (?, ?, ?, ?)", (key, time, None, 1))
        rusult.append(key)
    conn.commit()
    conn.close()
    return rusult


def check_sub_status(user_id):
    date_now = int(time())
    have_sub_time = None
    connect = sqlite3.connect('db.db')
    cursor = connect.cursor()
    cursor.execute("SELECT sub_time, date_get_sub FROM info WHERE user_id=?", [user_id])
    result = cursor.fetchone()
    connect.commit()
    if result:
        sub_time = result[0]
        date_get_sub = result[1]
        sub_status = sub_time - (date_now - date_get_sub)

        if sub_status > 0:
            cursor.execute("UPDATE info SET sub_status = ? WHERE user_id=?", [sub_status, user_id])
            connect.commit()
            have_sub_time = sub_status
        else:
            cursor.execute("UPDATE info SET sub_time = ?, sub_status = ? WHERE user_id=?",
                           [0, 0, user_id])
            connect.commit()
            have_sub_time = 0

    cursor.close()
    connect.close()
    return have_sub_time


def add_sub_info_to_logs(user_id, user_name, type_sub, sub_time, date_buy_sub):
    conn = sqlite3.connect('db.db')
    cursor = conn.cursor()
    cursor.execute("INSERT INTO subs_logi(user_id, user_name, type_sub, sub_time, date_buy_sub) VALUES(?, ?, ?, ?, ?)",
                   [user_id, user_name, type_sub, sub_time, date_buy_sub])
    conn.commit()
    conn.close()


def create_docs(QnumbSchet, Qdate, QbuyerData, QnumbContract, Qroute, QrouteDate, Qprice):
    wb = openpyxl.load_workbook('s.xlsx')
    ws = wb['TDSheet']

    nSchet = ws['B13']
    buyer = ws['F19']
    description = ws['D23']
    pricecell = ws['U23']
    textprice = ws['B30']

    numbSchet = str(QnumbSchet)
    date = str(Qdate)
    buyerData = str(QbuyerData)
    numbContract = str(QnumbContract)
    route = str(Qroute)
    routeDate = str(QrouteDate)
    price = int(Qprice)

    if numbContract == '-':
        numbContract = ''

    if route == '-':
        route = ''

    strPrice = num2words(price, lang='ru')
    strPrice = strPrice.capitalize()

    print('Счет готов!')

    nSchet.value = 'Счет на оплату №' + numbSchet + ' от ' + date + 'г.'
    buyer.value = buyerData
    description.value = 'Транспортно-экспедиционные услуги по договору-заявке ' + numbContract + ' ' + route + ' от ' + routeDate + 'г.'
    pricecell.value = str(price)
    textprice.value = strPrice + ' рублей 00 копеек'

    schet_filename = 'docs/счет ' + numbSchet + ' ' + str(price) + '.xlsx'
    wb.save(schet_filename)

    ### AKT ###

    wb2 = openpyxl.load_workbook('a.xlsx')
    ws2 = wb2['TDSheet']

    nAkts = ws2['B3']
    buyer2 = ws2['F7']
    description2 = ws2['D11']
    pricecell2 = ws2['Z11']
    textprice2 = ws2['B18']

    nAkts.value = 'Акт №' + numbSchet + ' от ' + date + 'г.'
    buyer2.value = buyerData
    description2.value = 'Транспортно-экспедиционные услуги по договору-заявке ' + numbContract + ' ' + route + ' от ' + routeDate + 'г.'
    pricecell2.value = str(price)
    textprice2.value = strPrice + ' рублей 00 копеек'

    akt_filename = 'docs/акт ' + numbSchet + ' ' + str(price) + '.xlsx'
    wb2.save(akt_filename)

    return schet_filename, akt_filename
