from aiogram.dispatcher.filters import Command
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.dispatcher import FSMContext
from aiogram.types import Message, ContentTypes, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from main import dp, bot
from data import config
from database import functions
from keyboards import keys
import datetime
import openpyxl

# sheet_index = "https://docs.google.com/spreadsheets/d/11Pp0kiyCf1o_IAiJf2NNbzV4"
#
#
# def connecting_gsheet(sheet_index=1) -> pygsheets.Spreadsheet:
#     gc = pygsheets.authorize(service_account_file="google/service-account.json")
#     return gc.open_by_key(config.SHEET_KEY)[sheet_index]
#
#
# def add_data_to_sheet(worksheet, values: list, next_empty_row: int, last_iteration: bool = False, financials=None) -> bool:
#     try:
#         if last_iteration and financials:
#             values.extend(financials)
#             range_to_update = f"A{next_empty_row}:N{next_empty_row}"
#         else:
#             range_to_update = f"A{next_empty_row}:J{next_empty_row}"
#
#         worksheet.update_values(range_to_update, [values])
#         return True
#     except Exception as e:
#         print(e)
#         return False


class GiftState(StatesGroup):
    WaitingForPhoneNumber = State()


class Review(StatesGroup):
    WaitingForReview = State()


def column_number_to_letter(column_number):
    """Преобразует числовой индекс столбца в буквенный индекс."""
    result = ""
    while column_number > 0:
        remainder = (column_number - 1) % 26
        result = chr(65 + remainder) + result
        column_number = (column_number - 1) // 26
    return result


@dp.message_handler(Command('start'))
async def start(message: Message):
    if functions.check_new_user(message.from_user.id):
        functions.add_user(message.from_user.id)
    else:
        pass

    keyboard = InlineKeyboardMarkup()
    keyboard.add(keys.ikb('Получить подарок', 'get_gift'), keys.ikb('Написать продавцу', 'support_request'))
    await bot.send_message(chat_id=message.chat.id,
                           text='Привет! Благодарим тебя за покупку!'
                                '\nНадеемся, что тебе или твоим близким понравится наш товар.  В качестве благодарности за доверие, мы хотим тебе сделать подарок 🎁',
                           reply_markup=keyboard, parse_mode='html')


@dp.message_handler(text='Написать продавцу')
async def support_request2(message: Message):
    keyboard = InlineKeyboardMarkup()
    keyboard.add(keys.ikb('Качество товара', 'ask_quality_product'),
                 keys.ikb('Комплектация товара', 'ask_equipment_product'),
                 keys.ikb('Другой вопрос', 'ask_any_question'))
    await message.answer(text='<b>Выбери пожалуйста категорию обращения: </b>', reply_markup=keyboard,
                         parse_mode='html')


@dp.message_handler(text='Получить подарок')
async def get_gift(message: Message):
    keyboard = keys.share_phone_markup()
    await bot.send_message(chat_id=message.from_user.id, text=f'{message.from_user.full_name}\n'
                                                              'У нас для тебя подарок за отзыв. \n\n'
                                                              'Пожалуйста, введи в ответном сообщении свой номер телефона, на него мы отправим бонус!\n'
                                                              'Или нажми на кнопку «Поделиться номером»',
                           reply_markup=keyboard)
    await GiftState.WaitingForPhoneNumber.set()


@dp.message_handler(text='Отсчет')
async def get_info_logs(message: Message):
    if message.from_user.id in config.ADMINS:
        logs = functions.get_info_logs()
        rows = functions.get_answers_logs()
        requests_value = logs[0]
        answers_value = logs[1]
        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet['A1'] = 'Дата'
        sheet['B1'] = 'Сообщение'
        sheet['C1'] = 'Пользователь'
        sheet['H1'] = "Всего обращений"
        sheet['I1'] = "Ответом менеджера"

        start_row = 2  # Начинаем с 2, так как 1 строка обычно используется для заголовков

        for index, value in enumerate((rows[0:][1][0]), start=2):
            sheet[f'A{index}'] = value
        for index, value in enumerate(rows[0:][1], start=2):
            sheet[f'B{index}'] = value
        for index, value in enumerate(rows[0:][2], start=2):
            sheet[f'C{index}'] = value

        sheet['I2'] = requests_value
        sheet['H2'] = answers_value

        # Проверяем условие и подсвечиваем ячейку
        if requests_value < answers_value:
            sheet['I2'].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000',
                                                           fill_type='solid')  # Красный
        elif requests_value == answers_value:
            sheet['I2'].fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00',
                                                           fill_type='solid')  # Зеленый
        # Сохранение файла
        wb.save('example.xlsx')

        with open('example.xlsx', 'rb') as excel_file:
            await message.reply_document(excel_file)



@dp.message_handler(text='Очистить таблицу')
async def clear_logs(message: Message):
    if message.from_user.id in config.ADMINS:
        pass


@dp.message_handler(Command('close'))
async def handle_message(message: Message):
    if message.from_user.id in config.ADMINS:
        command, *args = message.get_full_command()

        if args:
            try:
                user_id = int(args[0])
                keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
                keyboard.add(KeyboardButton('Поделиться номером', request_contact=True))
                await bot.send_message(chat_id=user_id,
                                       text=f'{message.from_user.full_name}, рады, что удалось решить ваш вопрос!\n'
                                            'У нас для тебя подарок за отзыв. \n\n'
                                            'Пожалуйста, введи в ответном сообщении свой номер телефона, на него мы отправим бонус!\n'
                                            'Или нажми на кнопку «Поделиться номером»',
                                       reply_markup=keyboard)
            except ValueError:
                await message.answer("ID пользователя должен быть числом.")
        else:
            await message.answer("Пожалуйста, введите ID пользователя после команды.")
    else:
        await message.answer("У вас нет прав на выполнение этой команды.")


@dp.message_handler(state=GiftState.WaitingForPhoneNumber, content_types=ContentTypes.TEXT)
async def handle_phone_number(message: Message, state: FSMContext):
    if message.text.isdigit() and len(message.text) == 11:
        phone_number = message.text
        print(phone_number)
        await message.answer(f"Ты ввел номер телефона: {phone_number}\n\n"
                             "Пожалуйста, подтверди, что номер верный:",
                             reply_markup=keys.get_confirmation_keyboard(phone_number))
        await state.finish()
    else:
        await message.answer("Пожалуйста, введи корректный номер телефона.")


@dp.message_handler(content_types=ContentTypes.CONTACT, state=GiftState.WaitingForPhoneNumber)
async def handle_contact1(message: Message):
    phone_number = message.contact.phone_number
    await message.answer(f"Ты поделился номером: {phone_number}\n\n"
                         "Пожалуйста, подтверди, что номер верный:",
                         reply_markup=keys.get_confirmation_keyboard(phone_number))


@dp.message_handler(content_types=ContentTypes.TEXT, state=Review.WaitingForReview)
async def handle_photo(message: Message):
    await message.answer('Нам нужен только скриншот.')


@dp.message_handler(content_types=ContentTypes.PHOTO, state=Review.WaitingForReview)
async def handle_photo(message: Message, state: FSMContext):
    await state.update_data(screenshot=message.photo[-1])
    photo = message.photo[-1]
    phone = functions.get_user_phone(message.from_user.id)
    await bot.send_photo(chat_id=config.GROUP_ID, photo=photo.file_id,
                         caption=f'{message.from_user.id}\nНужно проверить этот отзыв и начислить деньги на номер <b>{phone}</b>',
                         reply_markup=keys.get_feedback_confirmation_keyboard(), parse_mode='html')
    await message.answer('Ваш скриншот успешно отправлен на модерацию ✅', reply_markup=keys.main_markup())
    await state.finish()
